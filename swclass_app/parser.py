from __future__ import annotations

import json
import os
import tempfile
from pathlib import Path

from openpyxl import load_workbook


STOCK_CODE_COLUMN = "股票代码"
LEVEL1_INDUSTRY_COLUMN = "新版一级行业"
LEVEL2_INDUSTRY_COLUMN = "新版二级行业"
LEVEL3_INDUSTRY_COLUMN = "新版三级行业"


IndustryTree = list[dict[str, object]]


def parse_industry_stocks(xlsx_path: Path) -> IndustryTree:
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = next(rows)
        stock_idx = _column_index(headers, STOCK_CODE_COLUMN)
        level1_idx = _column_index(headers, LEVEL1_INDUSTRY_COLUMN)
        level2_idx = _column_index(headers, LEVEL2_INDUSTRY_COLUMN)
        level3_idx = _column_index(headers, LEVEL3_INDUSTRY_COLUMN)

        grouped: dict[str, dict[str, dict[str, set[str]]]] = {}
        for row in rows:
            stock_code = _clean_cell(row[stock_idx] if stock_idx < len(row) else None)
            level1 = _clean_cell(row[level1_idx] if level1_idx < len(row) else None)
            level2 = _clean_cell(row[level2_idx] if level2_idx < len(row) else None)
            level3 = _clean_cell(row[level3_idx] if level3_idx < len(row) else None)
            if not stock_code or not level1 or not level2 or not level3:
                continue
            grouped.setdefault(level1, {}).setdefault(level2, {}).setdefault(level3, set()).add(stock_code)

        return _format_industry_tree(grouped)
    finally:
        workbook.close()


def write_industry_json(data: IndustryTree, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fd, temp_name = tempfile.mkstemp(
        prefix=f".{output_path.name}.",
        suffix=".tmp",
        dir=output_path.parent,
        text=True,
    )
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as temp_file:
            json.dump(data, temp_file, ensure_ascii=False, indent=2)
            temp_file.write("\n")
        os.replace(temp_name, output_path)
    except Exception:
        Path(temp_name).unlink(missing_ok=True)
        raise


def _column_index(headers: tuple[object, ...], column_name: str) -> int:
    normalized = [_clean_cell(value) for value in headers]
    try:
        return normalized.index(column_name)
    except ValueError as exc:
        raise ValueError(f"未找到必要列: {column_name}") from exc


def _clean_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _format_industry_tree(grouped: dict[str, dict[str, dict[str, set[str]]]]) -> IndustryTree:
    return [
        {
            "name": level1,
            "children": [
                {
                    "name": level2,
                    "children": [
                        {"name": level3, "codes": sorted(codes)}
                        for level3, codes in sorted(level3_groups.items())
                    ],
                }
                for level2, level3_groups in sorted(level2_groups.items())
            ],
        }
        for level1, level2_groups in sorted(grouped.items())
    ]
